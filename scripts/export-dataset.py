#!/usr/bin/env python3
"""Export the Nihon place dataset from the master workbook to application JSON.

Canonical, reproducible replacement for the original export-dataset.mjs, which depended
on @oai/artifact-tool (not an installable package) and, for local editing, on LibreOffice
recalculation of the JPY->MXN price formulas. This script has neither dependency: it reads
the workbook with openpyxl and computes Precio MXN mín/máx itself from Precio JPY x the
Configuración!B4 exchange rate, deterministically, regardless of whether the workbook's
cached formula values are present.

Normalization rules (see docs/DATA_MODEL.md):
  - Preserve original Spanish editorial text in `raw` fields where parsing is uncertain.
  - Parse numeric coordinate, price, distance, and walking-minute columns as numbers.
  - Normalize Si/No reservation values to booleans while keeping the original value.
  - Convert duration text only when a safe range is clear; otherwise keep `raw` and leave
    numeric values empty. Durations counted in days or nights ("Dia completo", "1-2 dias")
    describe trip space, not time on site, so they keep only their raw text.
  - Do not infer an image URL from an image brief.
  - Keep official URLs and consultation dates attached to the relevant records.

Usage:
    python3 scripts/export-dataset.py <workbook.xlsx> [output-dir]

Output files (JSON, 2-space indent, UTF-8, no unicode escaping) written to output-dir
(default ./data): places.json, clusters.json, nearby.json, seasonal-alerts.json,
sources.json.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

SHEET_RANGES = {
    "places": ("Lugares", 215),
    "clusters": ("Clusters", 89),
    "nearby": ("Cercanos", 404),
    "seasonal-alerts": ("Feb-Mar 2027", 34),
    "sources": ("Fuentes", 40),
}

DAY_NIGHT_RE = re.compile(r"d[ií]as?\b|noches?\b", re.IGNORECASE)
RANGE_RE = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*[–—-]\s*(\d+(?:[.,]\d+)?)\s*(min|minutos?|h|horas?)",
    re.IGNORECASE,
)


def as_number(value):
    """Mirrors the JS asNumber(): comma-as-decimal-separator aware, NaN -> None."""
    if value is None:
        return None
    try:
        n = float(str(value).replace(",", "."))
    except ValueError:
        return None
    if n != n:  # NaN
        return None
    return int(n) if n.is_integer() else n


def parse_duration(raw):
    text = str(raw).strip() if raw is not None else ""
    if not text:
        return {"raw": None}
    if DAY_NIGHT_RE.search(text):
        return {"raw": text}
    match = RANGE_RE.search(text)
    if not match:
        return {"raw": text}
    multiplier = 60 if re.match(r"^h|hora", match.group(3), re.IGNORECASE) else 1
    return {
        "raw": text,
        "minMinutes": round(as_number(match.group(1)) * multiplier),
        "maxMinutes": round(as_number(match.group(2)) * multiplier),
    }


def sheet_objects(wb, sheet_name, max_row):
    """Reads header row + data rows [2, max_row] into a list of header-keyed dicts,
    skipping rows that are entirely empty — mirrors the JS rows()+objects() pair."""
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    result = []
    for row_num in range(2, max_row + 1):
        values = [ws.cell(row=row_num, column=c + 1).value for c in range(len(headers))]
        if any(v is not None and v != "" for v in values):
            result.append(dict(zip(headers, values)))
    return result


def normalize_place(row, mxn_rate):
    reservation_raw = str(row.get("Reserva") or "").strip()
    jpy_min = as_number(row.get("Precio JPY mín"))
    jpy_max = as_number(row.get("Precio JPY máx"))
    # Computed directly from JPY x Configuración!B4 rather than read from the workbook's
    # cached formula value, so this does not depend on the workbook having been recalculated.
    mxn_min = jpy_min * mxn_rate if jpy_min is not None else None
    mxn_max = jpy_max * mxn_rate if jpy_max is not None else None
    if mxn_min is not None and mxn_min == int(mxn_min):
        mxn_min = int(mxn_min)
    if mxn_max is not None and mxn_max == int(mxn_max):
        mxn_max = int(mxn_max)

    return {
        "id": row.get("ID"),
        "hub": row.get("Hub"),
        "region": row.get("Región"),
        "prefecture": row.get("Prefectura"),
        "municipality": row.get("Municipio"),
        "neighborhood": row.get("Zona/barrio"),
        "cluster": row.get("Cluster"),
        "name": row.get("Nombre oficial"),
        "japaneseName": row.get("Nombre japonés"),
        "mapTitle": row.get("Título mapa"),
        "category": row.get("Categoría"),
        "type": row.get("Tipo"),
        "grade": row.get("Grado"),
        "description": row.get("Descripción corta"),
        "differentiator": row.get("Qué lo hace diferente"),
        "experience": row.get("Qué se hace/ve"),
        "duration": parse_duration(row.get("Duración")),
        "bestTime": row.get("Mejor momento"),
        "bestSeason": row.get("Mejor época"),
        "crowdLevel": row.get("Aglomeración"),
        "tourismLevel": row.get("Qué tan turístico"),
        "price": {
            "currency": "JPY",
            "min": jpy_min,
            "max": jpy_max,
            "mxnMin": mxn_min,
            "mxnMax": mxn_max,
        },
        "reservation": {
            "required": reservation_raw.lower() == "sí",
            "leadTime": row.get("Anticipación"),
            "raw": reservation_raw,
        },
        "schedule": {
            "hours": row.get("Horario"),
            "closures": row.get("Días de cierre"),
        },
        "transport": row.get("Estación/transporte útil"),
        "accessibility": row.get("Accesibilidad aproximada"),
        "coordinates": {
            "lat": as_number(row.get("Latitud")),
            "lng": as_number(row.get("Longitud")),
        },
        "officialUrl": row.get("Página oficial/fuente"),
        "googleMapsUrl": row.get("Google Maps"),
        "imageBrief": row.get("Imagen recomendada"),
        "imageStatus": "brief-only",
        "nearbyIds": [],
        "hiddenGemStatus": row.get("Hidden Gem Status"),
        "alternativeTo": row.get("Alternativa a"),
        "updatedAt": row.get("Actualizado"),
        "febMar2027": {
            "status": row.get("Estado Feb–Mar 2027"),
            "warning": row.get("Advertencia / oportunidad Feb–Mar 2027"),
            "action": row.get("Acción recomendada Feb–Mar 2027"),
        },
    }


def export(workbook_path, output_dir):
    wb = openpyxl.load_workbook(workbook_path, data_only=False)

    mxn_rate = wb["Configuración"]["B4"].value
    if not isinstance(mxn_rate, (int, float)):
        raise SystemExit(f"Configuración!B4 (JPY -> MXN rate) is not numeric: {mxn_rate!r}")

    places_sheet, places_max = SHEET_RANGES["places"]
    places = [normalize_place(row, mxn_rate) for row in sheet_objects(wb, places_sheet, places_max)]

    collections = {"places": places}
    for name in ("clusters", "nearby", "seasonal-alerts", "sources"):
        sheet_name, max_row = SHEET_RANGES[name]
        collections[name] = sheet_objects(wb, sheet_name, max_row)

    nearby_by_place = {}
    for relation in collections["nearby"]:
        nearby_by_place.setdefault(relation["Desde ID"], []).append(relation["Hacia ID"])
    for place in places:
        place["nearbyIds"] = nearby_by_place.get(place["id"], [])

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    counts = {}
    for name, value in collections.items():
        with open(output_dir / f"{name}.json", "w", encoding="utf-8") as f:
            json.dump(value, f, indent=2, ensure_ascii=False)
        counts[name] = len(value)
    return counts


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    workbook_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "./data"
    counts = export(workbook_path, output_dir)
    print(json.dumps(counts))


if __name__ == "__main__":
    main()
